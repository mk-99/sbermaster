#!/usr/local/bin/python3

import imaplib, getpass, email, pprint, argparse, sys, ssl
import sberbank, vestabank

from email import policy as pl
from openpyxl import Workbook
from dateutil.parser import parse as date_parse

def process_mailbox(mb_reader, s=r"(SINCE 1-Mar-2017 FROM 900)", stop_words=lambda x: True):
    """
    Search messages from Sberbank in mailbox    
    :param mb_reader: mailbox reader object
    :param s: search string in IMAP format
    :return: messages as list of dicts {'time':headers, 'body':body}, time is datetime object, body is a string
    """

    sms_list = []

    print("Searching...")
    rv, data = mb_reader.search(None, s)
    if rv != 'OK':
        print("No messages found!")
        return sms_list

    mset = ""
    for d in data[0].split():
        mset = mset + "," + d.decode('ascii')

    print("Fetching...")
    rv, data = mb_reader.fetch(mset, '(RFC822)')
    if rv != 'OK':
        print("ERROR getting message", mset)
        return sms_list

    msg_list = []
    for m in data:
        if len(m) > 1:
            msg_list.append(m[1])

    print("Processing...")
    counter = 0
    for msg_bytes in msg_list:
        msg = email.message_from_bytes(msg_bytes, policy=pl.default)
        for part in msg.walk(): # Parse message
            sms_list.append({'time': date_parse(dict(part.items())['Date']), 'body': part.get_content()})
        counter += 1
        if counter % 100 == 0: # Make some awaiting progress
            print("Processed ", counter, "messages")

    return list(filter(stop_words, sms_list))


def process_arguments():
    """
    Processes command line arguments 
    :return: dict of configuration options
    """

    parser = argparse.ArgumentParser(description="Process Sberbank SMS messages backed up to imap server and generate"
                                                 "xlsx sheet",
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter
                                     )
    parser.add_argument("-d", "--date", help="Start from date", default="1-Mar-2017")
    parser.add_argument("-l", "--login", help="Login with this name")
    parser.add_argument("-p", "--password", help="Login with this password (prompt for password if none)")
    parser.add_argument("-s", "--site", help="Connect to this imap server", default="imap.gmail.com")
    parser.add_argument("-f", "--folder", help="Folder/label to read SMS from", default="SMS")
    parser.add_argument("-S", "--search", help="IMAP search string", default="FROM 900")
    parser.add_argument("-w", "--warn", help="Print warnings", action="store_true")
    parser.add_argument("-q", "--quiet", help="No print at all", action="store_true")
    parser.add_argument("-1", "--sms", help="Print SMS list and stop", action="store_true")
    parser.add_argument("-b", "--bank", help="'sberbank' or 'vesta' (also changes search string)", default="sberbank")
    parser.add_argument("outfile", help="Output MS Excel file, please add .xlsx explicitly")

    prog_arguments =  vars(parser.parse_args())

    if prog_arguments['bank'] == "vesta" and prog_arguments['search'] == "FROM 900":
        prog_arguments['search'] = "FROM VestaBank"

    return prog_arguments


def save_operations(arg, wb_file='sbercards.xlsx'):
    """
    Save operations to xlsx
    :param wb_file: write to this file
    :param arg: tuple of (oper, trf), oper is list of transactions as sms-es
            trf is list of transfers as sms-es
    :return: None
    """

    oper, trf = arg

    wb = Workbook()
    ws = wb.active
    ws.title = "Operations"

    i = 1
    for val in ("Card", "Time", "Time inside SMS", "Operation", "Sum", "Currency",
                "Comission", "Comm. currency", "Balance", "Place",
                "Name", "Comment", "Time of transfer"
               ):
        ws.cell(row=1, column=i, value=val)
        i += 1

    cur_row = 2
    for o in oper:
        ws.cell(row=cur_row, column=1, value=o['card'])
        ws.cell(row=cur_row, column=2, value=o['time'])
        ws.cell(row=cur_row, column=3, value=o['time1'])
        ws.cell(row=cur_row, column=4, value=o['oper'])
        ws.cell(row=cur_row, column=5, value=o['sum'])
        ws.cell(row=cur_row, column=6, value=o['currency'])
        ws.cell(row=cur_row, column=7, value=o['comission'])
        ws.cell(row=cur_row, column=8, value=o['commcurr'])
        ws.cell(row=cur_row, column=9, value=o['bal'])
        ws.cell(row=cur_row, column=10, value=o['place'])
        if ('transfer' not in o.keys()) or ('transfer' in o.keys() and o['transfer'] == 'Not found'):
            ws.cell(row=cur_row, column=11, value="")
            ws.cell(row=cur_row, column=12, value="")
            ws.cell(row=cur_row, column=13, value="")
        else:
            ws.cell(row=cur_row, column=11, value=o['transfer']['name'])
            ws.cell(row=cur_row, column=12, value=o['transfer']['comment'])
            ws.cell(row=cur_row, column=13, value=o['transfer']['time'])
        cur_row += 1

    ws1 = wb.create_sheet("Transfers")

    i = 1
    for val in ('Time', 'Name', 'Sum', 'Comment'):
        ws1.cell(row=1, column=i, value=val)
        i += 1

    cur_row = 2
    for transaction in trf:
        ws1.cell(row=cur_row, column=1, value=transaction['time'])
        ws1.cell(row=cur_row, column=2, value=transaction['name'])
        ws1.cell(row=cur_row, column=3, value=transaction['sum'])
        ws1.cell(row=cur_row, column=4, value=transaction['comment'])
        cur_row += 1

    wb.save(wb_file)

if __name__ == "__main__":

    config_opts = process_arguments()

    if config_opts['bank'] == "vesta":
        process_sms_list = vestabank.process_sms_list
        stop_words = vestabank.stop_words
    elif config_opts['bank'] == "sberbank":
        process_sms_list = sberbank.process_sms_list
        stop_words = sberbank.stop_words
    else:
        print("ERROR: unknown bank ", config_opts['bank'])
        sys.exit(1)

    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE
    mb_reader = imaplib.IMAP4_SSL(config_opts['site'], ssl_context=context)

    try:
        rv, data = mb_reader.login(config_opts['login'],
                                   config_opts['password'] if config_opts['password'] else getpass.getpass()
                                   )
    except imaplib.IMAP4.error:
        print("LOGIN FAILED!!! ")
        sys.exit(1)

    rv, data = mb_reader.select(config_opts['folder'])
    if rv == 'OK':
        search_string = "(" + config_opts['search'] + " SINCE " + config_opts['date'] + ")"
        sms_list = process_mailbox(mb_reader, search_string, stop_words)
        mb_reader.close()
    else:
        print("ERROR: Unable to open mailbox ", rv)
        sys.exit(1)

    if sms_list:
        if config_opts['sms']:
            for sms in sms_list:
                print(sms['body'])
        else:
            save_operations((process_sms_list(sms_list, warn=config_opts['warn'])), wb_file=config_opts['outfile'])

    mb_reader.logout()
